from io import BytesIO
from datetime import date
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _pdf_base(doc, employee_name, start, end):
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleX',
        parent=styles['Title'],
        fontSize=18,
        textColor=colors.HexColor('#1e293b'),
        spaceAfter=2,
    )
    sub_style = ParagraphStyle(
        'SubX',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#64748b'),
        alignment=TA_CENTER,
        spaceAfter=8,
    )

    doc.append(Paragraph('Relatório de Diárias - Limpeza Airbnb', title_style))
    doc.append(Paragraph(
        f'{employee_name} &nbsp;|&nbsp; Período: {start} a {end}', sub_style
    ))
    doc.append(Spacer(1, 6))


def employee_report_pdf(data, employee_name: str, start: date, end: date) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )

    story = []
    _pdf_base(story, employee_name, start, end)

    # Resumo
    summary = [
        ['Dias trabalhados', 'Diárias inteiras', 'Meias diárias', 'Tarefas', 'Concluídas', 'Taxa'],
        [
            str(data['total_days_worked']),
            str(data['full_day_count']),
            str(data['half_day_count']),
            str(data['total_tasks']),
            str(data['completed_tasks']),
            f"{data['completion_rate']}%",
        ],
    ]
    summary_table = Table(summary, colWidths=[28 * mm] * 6)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, 1), [colors.HexColor('#f5f3ff'), colors.HexColor('#ffffff')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 14))

    # Tarefas
    rows = [['Data', 'Apartamento', 'Tipo', 'Status']]
    for t in data['tasks']:
        rows.append([
            str(t['scheduled_date']),
            t['apartment_name'],
            t['task_type'],
            t['status'],
        ])

    tasks_table = Table(rows, colWidths=[26 * mm, 72 * mm, 34 * mm, 30 * mm])
    tasks_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(tasks_table)

    doc.build(story)
    buffer.seek(0)
    return buffer


def general_report_pdf(data, start: date, end: date) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )

    story = []
    _pdf_base(story, 'Relatório Geral', start, end)

    summary = [
        ['Funcionários', 'Tarefas', 'Concluídas', 'Taxa de conclusão'],
        [
            str(data['total_employees']),
            str(data['total_tasks']),
            str(data['completed_tasks']),
            f"{data['completion_rate']}%",
        ],
    ]
    summary_table = Table(summary, colWidths=[42 * mm] * 4)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, 1), [colors.HexColor('#f5f3ff'), colors.HexColor('#ffffff')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 14))

    rows = [['Funcionário', 'Diárias Inteiras', 'Meias Diárias', 'Tarefas', 'Concluídas', 'Conclusão']]
    for e in data['employees']:
        rate = (
            f"{round((e['completed_tasks'] / e['total_tasks']) * 100)}%"
            if e['total_tasks'] > 0 else '-'
        )
        rows.append([
            e['employee_name'],
            str(e['full_day_count']),
            str(e['half_day_count']),
            str(e['total_tasks']),
            str(e['completed_tasks']),
            rate,
        ])

    tasks_table = Table(rows, colWidths=[58 * mm, 26 * mm, 24 * mm, 20 * mm, 20 * mm, 20 * mm])
    tasks_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(tasks_table)

    doc.build(story)
    buffer.seek(0)
    return buffer


def _excel_header(ws, headers, start_row=1):
    ws.cell(row=start_row, column=1, value='Limpeza Airbnb - Relatório de Diárias')
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=14, color='1E293B')
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))

    fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row + 1, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = fill
        c.alignment = Alignment(horizontal='center')


def employee_report_xlsx(data, employee_name: str, start: date, end: date) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Relatório'

    _excel_header(ws, ['Indicador', 'Valor'])
    ws.append(['Funcionário', employee_name])
    ws.append(['Período', f'{start} a {end}'])
    ws.append(['Dias trabalhados', data['total_days_worked']])
    ws.append(['Diárias inteiras', data['full_day_count']])
    ws.append(['Meias diárias', data['half_day_count']])
    ws.append(['Total de tarefas', data['total_tasks']])
    ws.append(['Concluídas', data['completed_tasks']])
    ws.append(['Taxa de conclusão', f"{data['completion_rate']}%"])

    # Segunda planilha com tarefas
    ws2 = wb.create_sheet('Tarefas')
    _excel_header(ws2, ['Data', 'Apartamento', 'Tipo', 'Status'])
    for t in data['tasks']:
        ws2.append([str(t['scheduled_date']), t['apartment_name'], t['task_type'], t['status']])

    for wsx in (ws, ws2):
        for col_cells in wsx.columns:
            max_len = max((len(str(c.value or '')) for c in col_cells), default=0)
            wsx.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 4

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def general_report_xlsx(data, start: date, end: date) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Relatório Geral'

    _excel_header(ws, ['Funcionário', 'Diárias Inteiras', 'Meias Diárias', 'Tarefas', 'Concluídas', 'Conclusão'])
    for e in data['employees']:
        rate = (
            f"{round((e['completed_tasks'] / e['total_tasks']) * 100)}%"
            if e['total_tasks'] > 0 else '-'
        )
        ws.append([
            e['employee_name'],
            e['full_day_count'],
            e['half_day_count'],
            e['total_tasks'],
            e['completed_tasks'],
            rate,
        ])

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 4

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer